import os
import sys
import openpyxl
import datetime

# We assume all files are in a correct format
# No validation will happen here
def combine_xlsx(files, outfile = None):
	# Store combined information
	headers = None
	items = []

	# Loop over given files
	for f in files:
		# Open XLSX file and select active sheet as main
		workbook = openpyxl.load_workbook(f)
		sheet = workbook.active

		# Minimum row to start at.
		# If we already got the headers, start at 2
		min = 1
		if headers != None:
			min += 1

		# Loop over all rows
		for row in sheet.iter_rows(min_row=min):
			# Store upcoming row
			item = []
			for cell in row:
				item.append(cell.value)
			
			# Set headers if empty
			if headers == None:
				headers = item
			else:
				items.append(item)
	
	# Sort all data on column 2 (3rtd column, Name)
	items = sorted(items, key=lambda x: (x[2]))

	# Create new XLSX file
	workbook = openpyxl.Workbook()
	sheet = workbook.active

	# Write headers
	for i in range(1, len(headers)):
		sheet.cell(column=i, row=1, value=headers[i - 1])

	# Write data sheet
	for item in range(len(items)):
		for col in range(1, len(items[item]) + 1):
			sheet.cell(column=col, row=(item + 2), value=items[item][col - 1])

	# Set filename and save
	if outfile == None:
		outfile = "combined_{}.xlsx".format(datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
	workbook.save(outfile)

# Use as standalone
if __name__ == '__main__':
	# Get all files from the commandline.
	# Remove the first, since that's the executable name
	files = sys.argv
	files.pop(0)

	# Need minimum of 2 files
	if (len(files) < 2):
		print("Please provide at least 2 files.")
		exit(1)

	# Check extension
	extension = None
	for arg in sys.argv:
		ext = os.path.splitext(arg)[1].lower()
		if (extension == None):
			extension = ext
		elif (extension != ext):
			print("Please provide files of the same extension.")
			exit(1)
	
	# Check if all files exist
	for f in files:
		if os.path.exists(f) == False or os.path.isfile(f) == False:
			print("Path is not a valid file: {}".format(f))
			exit(1)
	
	# Switch to correct combiner
	if extension == ".xlsx":
		combine_xlsx(files)